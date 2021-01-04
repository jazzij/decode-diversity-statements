"""
This script will convert rows of a XLXS file into a series of text documents.
The title of the document will be name of org, the text of the document taken from the text column.

"""
from openpyxl import load_workbook
import os
# filename = "Sci_Tech_BLM_List_wText_Organizations.xlsx"
# delimiter = '\t'
# statements = []

# create directory to store generated text files (if one does not already exist)
curdir = os.getcwd()
textdir = curdir + "/texts"
if not os.path.isdir(textdir):
	os.mkdir(textdir)

#Get appropriate sheet from xcel file
#wb = load_workbook("Sci_Tech_BLM_List_wText_6-23.xlsx")
#statements = wb['Text-Only']
wb = load_workbook("Sci_Tech_BLM_List_wText.xlsx")
statements = wb['Sheet6']

# Save contexts of text into separate txt files
for row in statements.values:
	name, text = row
	if name is None:
		continue
	
	# create new filename. make sure good punctuation
	name = name.replace('/', '_')
	name = name.replace('*', '')
	name = name.replace(',', '')
	name = name.replace('. ', '_')
	txtfilename = name + ".txt"		
	path_to_file =os.path.join(textdir, txtfilename)
	
	# if file already exists, skip
	if os.path.isfile(path_to_file):
		continue
	
	# write the file
	file = open( path_to_file, 'w')
	if text:
		file.write(text)
	else:
		print("Row {} returned bad value {}".format(name, text))
	file.close()




